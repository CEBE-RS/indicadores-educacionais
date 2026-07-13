# -*- coding: utf-8 -*-
"""
Gera planilhas .xlsx formatadas (Central de Dados Abertos) a partir dos JSONs
do painel. Uma planilha por tema; abas separadas por rede.

Saida: painel/dados/downloads/*.xlsx  +  manifest.json (para o frontend)

Fonte dos dados: INEP (Censo Escolar, Taxas de Rendimento, IDEB/SAEB, INSE) e SAERS/RS.
"""
import sys, io, os, json, glob


# --- caminhos portateis (repo Git + bases locais) ---
from paths import BASE, OUT_DIR, PAINEL_DIR, BASES_DIR, BASES_BASICAS  # noqa: E402

DADOS = OUT_DIR
OUT = os.path.join(DADOS, "downloads")
os.makedirs(OUT, exist_ok=True)

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ?? Estilo
PRI = "0D47A1"; GOLD = "D4A84B"; HEADER_FILL = PatternFill("solid", fgColor=PRI)
ZEBRA = PatternFill("solid", fgColor="EEF3FA")
TITLE_FILL = PatternFill("solid", fgColor=PRI)
SUB_FILL = PatternFill("solid", fgColor="E8EDF4")
WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
SUB_FONT = Font(name="Calibri", bold=True, color=PRI, size=11)
SRC_FONT = Font(name="Calibri", italic=True, color="6B7280", size=9)
CELL_FONT = Font(name="Calibri", size=10)
thin = Side(style="thin", color="D6DEE8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

REDES = [("estadual", "Estadual"), ("municipal", "Municipal"),
         ("federal", "Federal"), ("privada", "Privada"), ("todas", "Todas")]

SKIP_KEYS = {"por_turno", "por_serie", "fonte", "dist_niveis_escolas"}

def flatten(obj, prefix=""):
    out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in SKIP_KEYS:
                continue
            nk = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, dict):
                out.update(flatten(v, nk))
            elif isinstance(v, list):
                continue
            else:
                out[nk] = v
    else:
        out[prefix] = obj
    return out

def ordered_cols(rows):
    cols = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k); cols.append(k)
    return cols

# ?? Helpers de escrita
def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            ln = len(str(cell.value))
            widths[col] = min(max(widths.get(col, 8), ln + 2), 44)
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_table(ws, start_row, headers, rows, freeze=False):
    r = start_row
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.fill = HEADER_FILL; c.font = WHITE_BOLD; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if freeze:
        ws.freeze_panes = ws.cell(row=r + 1, column=2)
    r += 1
    for i, row in enumerate(rows):
        for j, h in enumerate(headers, 1):
            v = row.get(h)
            if isinstance(v, float):
                v = round(v, 2)
            c = ws.cell(row=r, column=j, value=v)
            c.font = CELL_FONT; c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j > 2 else "left")
            if i % 2 == 1:
                c.fill = ZEBRA
        r += 1
    return r

def write_simple_sheet(wb, title, headers, rows):
    """Cria uma aba com tabela convencional: cabecalho na linha 1, dados abaixo,
    com linha congelada e autofiltro."""
    ws = wb.create_sheet(title=title[:31])
    if not headers:
        return ws
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL; c.font = WHITE_BOLD; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(rows):
        for j, h in enumerate(headers, 1):
            v = row.get(h)
            if isinstance(v, float):
                v = round(v, 2)
            c = ws.cell(row=i + 2, column=j, value=v)
            c.font = CELL_FONT; c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j > 2 else "left")
            if i % 2 == 1:
                c.fill = ZEBRA
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    autosize(ws)
    return ws

def write_title(ws, row, text, ncols, fill=TITLE_FILL, font=TITLE_FONT, h=24):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill; c.font = font
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = h
    return row + 1

def write_source(ws, row, fonte, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
    c = ws.cell(row=row, column=1, value="Fonte: " + fonte)
    c.font = SRC_FONT
    return row + 1

# ?? Extracao por dataset
def serie_rows(data, serie_key, ano_label="Ano"):
    serie = data.get(serie_key, {})
    rows = []
    for ano in sorted(serie.keys()):
        rec = {ano_label: ano}
        rec.update(flatten(serie[ano]))
        rows.append(rec)
    return rows

def mun_rows(data, mun_key, lookup_key="lookup_municipios", year=None, has_year=True):
    """Extrai linhas por municipio. Com has_year=True e year=None, exporta todos os anos."""
    mun = data.get(mun_key, {})
    lookup = data.get(lookup_key, {})
    rows = []

    if has_year:
        if not mun:
            return [], None
        anos = [year] if year else sorted(mun.keys())
        for ano in anos:
            block = mun.get(ano, {})
            for cod in sorted(block.keys()):
                rec = {"Ano": ano, "Codigo IBGE": cod, "Municipio": lookup.get(cod, cod)}
                rec.update(flatten(block[cod]))
                rows.append(rec)
        if year:
            mano = str(year)
        elif len(anos) == 1:
            mano = str(anos[0])
        else:
            mano = f"{anos[0]}-{anos[-1]}"
    else:
        ano = year or "—"
        block = mun
        for cod in sorted(block.keys()):
            rec = {"Ano": ano, "Codigo IBGE": cod, "Municipio": lookup.get(cod, cod)}
            rec.update(flatten(block[cod]))
            rows.append(rec)
        mano = str(ano) if ano != "—" else None

    return rows, mano

def load(rede, file_tmpl):
    path = os.path.join(DADOS, file_tmpl.format(rede=rede))
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)

# ?? Definicao dos datasets/temas
def make_infra_transform(data):
    labels = data.get("labels", {})
    def t(h):
        if h.startswith("indicadores.") and h.endswith(".pct"):
            code = h.split(".")[1]
            return (labels.get(code, code)) + " (%)"
        if h == "total_escolas":
            return "Total de escolas"
        return h
    return t

def infra_keep(h):
    if h in ("Ano", "Codigo IBGE", "Municipio", "total_escolas", "escolas"):
        return True
    if h.startswith("indicadores.") and h.endswith(".pct"):
        return True
    return False

DATASETS = {
    "acesso": dict(short="Acesso", file="4_1_acesso_{rede}.json", serie="serie_temporal",
                   mun="por_municipio", has_year=True, primary=True),
    "fluxo": dict(short="Fluxo", file="4_3_fluxo_{rede}.json", serie="serie_temporal",
                  mun="por_municipio", has_year=True, primary=True),
    "tdi": dict(short="TDI", file="4_10_tdi_{rede}.json", serie="serie_temporal",
                mun="por_municipio", has_year=True, primary=True),
    "infra": dict(short="Infra", file="4_5_infra_{rede}.json", serie="serie_temporal",
                  mun="por_municipio", has_year=True, primary=True,
                  transform=make_infra_transform, keep=infra_keep),
    "docentes": dict(short="Docentes", file="4_5_docentes_{rede}.json", serie="serie_temporal_total",
                     mun="por_municipio_2025", has_year=False, mun_year="2025", primary=True),
    "afd": dict(short="AFD", file="4_9_afd_{rede}.json", serie="serie_temporal",
                mun="por_municipio", has_year=True, primary=True),
    "icg": dict(short="ICG", file="4_8_icg_{rede}.json", serie="serie_temporal",
                mun="por_municipio", has_year=True, primary=True),
    "ideb": dict(short="IDEB", file="4_7_ideb_{rede}.json", serie="serie_temporal",
                 mun="por_municipio", has_year=True, primary=True),
    "saeb": dict(short="SAEB", file="4_6_saeb_{rede}.json", serie="serie_temporal",
                 mun="por_municipio", has_year=True, primary=True),
    "inse": dict(short="INSE", file="4_7_inse_{rede}.json", serie="serie_temporal",
                 mun="por_municipio", has_year=True, primary=True),
}

# Um card por SECAO do painel (espelha a navegacao do dashboard).
THEMES = [
    dict(id="acesso", nome="Acesso e Matriculas", arquivo="dados_acesso_matriculas.xlsx",
         fonte="INEP - Censo Escolar (Microdados)", datasets=["acesso"], funil=True,
         desc="Matriculas por etapa, dependencia, localizacao, perfil dos alunos e funil de fluxo."),
    dict(id="infra", nome="Infraestrutura", arquivo="dados_infraestrutura.xlsx",
         fonte="INEP - Censo Escolar (Microdados de Escolas)", datasets=["infra"],
         desc="Indicadores de infraestrutura escolar (tecnologia, espacos, acessibilidade, saneamento)."),
    dict(id="icg", nome="Complexidade de Gestao", arquivo="dados_complexidade_gestao.xlsx",
         fonte="INEP - Indicador de Complexidade de Gestao (ICG)", datasets=["icg"],
         desc="Indicador de Complexidade de Gestao das escolas por nivel e por municipio."),
    dict(id="inse", nome="Contexto Socioeconomico (INSE)", arquivo="dados_inse.xlsx",
         fonte="INEP - Nivel Socioeconomico (INSE)", datasets=["inse"],
         desc="Nivel socioeconomico medio dos estudantes (INSE) por rede e por municipio."),
    dict(id="docencia", nome="Docencia", arquivo="dados_docencia.xlsx",
         fonte="INEP - Censo Escolar (Microdados de Docentes)", datasets=["docentes"],
         desc="Perfil docente, escolaridade, vinculo, faixa etaria e razao aluno-professor."),
    dict(id="afd", nome="Formacao Docente (AFD)", arquivo="dados_afd.xlsx",
         fonte="INEP - Adequacao da Formacao Docente (AFD)", datasets=["afd"],
         desc="Adequacao da formacao docente (AFD) por etapa de ensino e por municipio."),
    dict(id="fluxo", nome="Fluxo e Rendimento", arquivo="dados_fluxo_rendimento.xlsx",
         fonte="INEP - Indicadores Educacionais (Taxas de Rendimento)", datasets=["fluxo"],
         desc="Taxas de aprovacao, reprovacao e abandono por etapa e por serie."),
    dict(id="tdi", nome="Distorcao Idade-Serie", arquivo="dados_tdi.xlsx",
         fonte="INEP - Distorcao Idade-Serie (TDI)", datasets=["tdi"],
         desc="Taxa de distorcao idade-serie (TDI) por etapa de ensino e por municipio."),
    dict(id="saers", nome="SAERS", arquivo="dados_saers.xlsx",
         fonte="SEDUC-RS / CAED - SAERS", datasets=[], saers=True,
         desc="Proficiencias medias e padroes de desempenho do SAERS/RS por etapa e disciplina."),
    dict(id="desigualdades", nome="Desigualdades", arquivo="dados_desigualdades.xlsx",
         fonte="SEDUC-RS / CAED - SAERS (Microdados)", datasets=[], desig=True,
         desc="Resultados do SAERS por recortes de equidade (sexo, raca, deficiencia, localizacao, turno)."),
    dict(id="saeb", nome="SAEB", arquivo="dados_saeb.xlsx",
         fonte="INEP - SAEB", datasets=["saeb"],
         desc="Proficiencias do SAEB em Lingua Portuguesa e Matematica por rede, etapa e municipio."),
    dict(id="ideb", nome="IDEB", arquivo="dados_ideb.xlsx",
         fonte="INEP - IDEB", datasets=["ideb"],
         desc="IDEB observado e metas por rede, etapa de ensino e por municipio."),
]

# ?? Construcao das abas
def _disp(rows, keep, transform):
    cols = ordered_cols(rows)
    if keep:
        cols = [c for c in cols if keep(c)]
    headers = [transform(c) for c in cols]
    disp = [{transform(c): row.get(c) for c in cols} for row in rows]
    return headers, disp

def build_dataset_sheets(wb, ds_key, index_entries):
    ds = DATASETS[ds_key]
    transform_factory = ds.get("transform")
    keep = ds.get("keep")
    for rede, rede_label in REDES:
        data = load(rede, ds["file"])
        if not data:
            continue
        transform = transform_factory(data) if transform_factory else (lambda h: h)

        srows = serie_rows(data, ds["serie"])
        mrows, mano = ([], None)
        # Base por municipio apenas para o dataset principal de cada tema
        # (evita excesso de abas; secundarios trazem so a serie historica).
        if ds.get("mun") and ds.get("primary"):
            mrows, mano = mun_rows(data, ds["mun"], year=ds.get("mun_year"), has_year=ds.get("has_year", True))

        if srows:
            headers, disp = _disp(srows, keep, transform)
            title = f"{ds['short']} {rede_label}"[:31]
            write_simple_sheet(wb, title, headers, disp)
            index_entries.append((title, f"{ds['short']} - serie historica ({rede_label})"))

        if mrows:
            headers, disp = _disp(mrows, keep, transform)
            title = f"{ds['short']} {rede_label} Mun"[:31]
            write_simple_sheet(wb, title, headers, disp)
            index_entries.append((title, f"{ds['short']} - por municipio, serie historica {mano or ''} ({rede_label})"))

def build_funil_sheets(wb, index_entries):
    path = os.path.join(DADOS, "4_1_funil_turma_locdif.json")
    if not os.path.exists(path):
        return
    data = json.load(open(path, encoding="utf-8"))
    blocks = [("funil_por_serie", "Funil por Serie"),
              ("tamanho_turma", "Tamanho de Turma"),
              ("localizacao_diferenciada", "Localizacao Diferenciada")]
    for key, label in blocks:
        d = data.get(key)
        if not d:
            continue
        rows = []
        for ano in sorted(d.keys()):
            rec = {"Ano": ano}
            rec.update(flatten(d[ano]))
            rows.append(rec)
        if not rows:
            continue
        headers = ordered_cols(rows)
        write_simple_sheet(wb, label, headers, rows)
        index_entries.append((label[:31], label))

_DESIG_DIM_LABELS = {
    "deficiencia": "Deficiencia", "localizacao": "Localizacao", "turno": "Turno",
    "sexo": "Sexo", "raca": "Raca/Cor", "raca_loc": "Raca x Localizacao",
    "raca_sexo": "Raca x Sexo",
}
_DESIG_DISC_LABELS = {"LP": "Lingua Portuguesa", "MT": "Matematica"}

def _desig_metrics(rec):
    p = rec.get("padrao", {}) or {}
    return {
        "Avaliados": rec.get("n_padrao") or rec.get("n"),
        "Proficiencia media": rec.get("media"),
        "% Adequado+Avancado": rec.get("pct_adeq_av"),
        "% Basico": p.get("pct_basico"),
        "% Abaixo do basico": p.get("pct_abaixo"),
    }

def _desig_split(ed):
    etapa, disc = ed.rsplit("_", 1)
    return etapa, disc

def build_desig_sheets(wb, index_entries):
    path = os.path.join(DADOS, "4_11_desigualdades.json")
    if not os.path.exists(path):
        return
    data = json.load(open(path, encoding="utf-8"))
    anos = data.get("anos", [])
    if not anos:
        return
    etapa_labels = (data.get("metadata", {}) or {}).get("etapa_labels", {})
    # Nome dos municipios a partir de um dataset que tenha o lookup
    mun_lookup = {}
    acesso = load("estadual", "4_1_acesso_{rede}.json")
    if acesso:
        mun_lookup = acesso.get("lookup_municipios", {})

    def etapa_nome(e):
        return etapa_labels.get(e, e)

    # Aba 1: Geral (Estado) - serie por ano/etapa/disciplina
    grows = []
    for a in anos:
        ano = a.get("ano")
        for ed, rec in (a.get("geral", {}) or {}).items():
            etapa, disc = _desig_split(ed)
            row = {"Ano": ano, "Etapa": etapa_nome(etapa),
                   "Disciplina": _DESIG_DISC_LABELS.get(disc, disc)}
            row.update(_desig_metrics(rec))
            grows.append(row)
    if grows:
        headers = ordered_cols(grows)
        write_simple_sheet(wb, "Geral (Estado)", headers, grows)
        index_entries.append(("Geral (Estado)", "SAERS - proficiencia geral por ano, etapa e disciplina"))

    # Aba 2: Por grupo (Estado) - recortes de equidade
    drows = []
    for a in anos:
        ano = a.get("ano")
        for dim, grupos in (a.get("dimensoes", {}) or {}).items():
            for grupo, eds in (grupos or {}).items():
                for ed, rec in (eds or {}).items():
                    etapa, disc = _desig_split(ed)
                    row = {"Ano": ano, "Recorte": _DESIG_DIM_LABELS.get(dim, dim),
                           "Grupo": grupo, "Etapa": etapa_nome(etapa),
                           "Disciplina": _DESIG_DISC_LABELS.get(disc, disc)}
                    row.update(_desig_metrics(rec))
                    drows.append(row)
    if drows:
        headers = ordered_cols(drows)
        write_simple_sheet(wb, "Por Grupo (Estado)", headers, drows)
        index_entries.append(("Por Grupo (Estado)", "SAERS - recortes por sexo, raca, deficiencia, localizacao e turno"))

    # Aba 3: Por municipio — serie historica (todos os anos disponiveis)
    mrows = []
    for a in anos:
        ano = a.get("ano")
        pm = a.get("por_municipio", {}) or {}
        for cod in sorted(pm.keys()):
            geral = (pm[cod] or {}).get("geral", {}) or {}
            for ed, rec in geral.items():
                etapa, disc = _desig_split(ed)
                row = {"Ano": ano, "Codigo IBGE": cod, "Municipio": mun_lookup.get(cod, cod),
                       "Etapa": etapa_nome(etapa), "Disciplina": _DESIG_DISC_LABELS.get(disc, disc)}
                row.update(_desig_metrics(rec))
                mrows.append(row)
    if mrows:
        headers = ordered_cols(mrows)
        write_simple_sheet(wb, "Por Municipio", headers, mrows)
        anos_list = [a.get("ano") for a in anos if a.get("ano")]
        arange = f"{anos_list[0]}-{anos_list[-1]}" if len(anos_list) > 1 else str(anos_list[0] if anos_list else "")
        index_entries.append(("Por Municipio", f"SAERS - proficiencia geral por municipio, serie historica ({arange})"))

def build_saers_sheets(wb, index_entries):
    for rede, rede_label in [("estadual", "Estadual"), ("municipal", "Municipal"), ("todas", "Todas")]:
        path = os.path.join(DADOS, f"4_saers_{rede}.json")
        if not os.path.exists(path):
            continue
        data = json.load(open(path, encoding="utf-8"))
        anos = data.get("anos", [])
        if not anos:
            continue
        srows = []
        for a in anos:
            rec = {"Ano": a.get("ano"), "Total alunos": a.get("total_alunos"),
                   "Total avaliados": a.get("total_avaliados")}
            rec.update(flatten(a.get("geral", {})))
            srows.append(rec)
        headers = ordered_cols(srows)
        title = f"SAERS {rede_label}"[:31]
        write_simple_sheet(wb, title, headers, srows)
        index_entries.append((title, f"SAERS - serie historica ({rede_label})"))

def build_index_sheet(wb, theme, index_entries):
    ws = wb.create_sheet(title="Sobre", index=0)
    ws.sheet_view.showGridLines = False
    r = write_title(ws, 1, f"Central de Dados Abertos  -  {theme['nome']}", 4)
    r += 1
    meta = [("Tema", theme["nome"]), ("Descricao", theme["desc"]),
            ("Fonte", theme["fonte"]), ("Recorte", "Rio Grande do Sul (497 municipios, 30 CREs)")]
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color=PRI)
        c = ws.cell(row=r, column=2, value=v); c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    r += 1
    r = write_title(ws, r, "Abas desta planilha", 4, fill=SUB_FILL, font=SUB_FONT, h=18)
    ws.cell(row=r, column=1, value="Aba").fill = HEADER_FILL
    ws.cell(row=r, column=1).font = WHITE_BOLD
    ws.cell(row=r, column=2, value="Conteudo").fill = HEADER_FILL
    ws.cell(row=r, column=2).font = WHITE_BOLD
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    for aba, cont in index_entries:
        ws.cell(row=r, column=1, value=aba).font = CELL_FONT
        c = ws.cell(row=r, column=2, value=cont); c.font = CELL_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    ws.column_dimensions["A"].width = 26
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16

def main():
    manifest = []
    for theme in THEMES:
        print(f"\n== {theme['nome']} ==")
        wb = Workbook()
        wb.remove(wb.active)
        index_entries = []
        for ds_key in theme["datasets"]:
            build_dataset_sheets(wb, ds_key, index_entries)
        if theme.get("funil"):
            build_funil_sheets(wb, index_entries)
        if theme.get("saers"):
            build_saers_sheets(wb, index_entries)
        if theme.get("desig"):
            build_desig_sheets(wb, index_entries)
        build_index_sheet(wb, theme, index_entries)
        out_path = os.path.join(OUT, theme["arquivo"])
        wb.save(out_path)
        size_kb = round(os.path.getsize(out_path) / 1024)
        n_abas = len(index_entries)
        print(f"   -> {theme['arquivo']} ({size_kb} KB, {n_abas} abas de dados)")
        manifest.append({
            "id": theme["id"], "nome": theme["nome"], "arquivo": theme["arquivo"],
            "descricao": theme["desc"], "fonte": theme["fonte"],
            "abas": n_abas, "tamanho_kb": size_kb,
        })
    with open(os.path.join(OUT, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"\nManifest: {len(manifest)} temas -> downloads/manifest.json")

if __name__ == "__main__":
    main()
